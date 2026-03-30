from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.utils import timezone
import datetime
# openpyxl imports moved inside method to avoid startup crash if not installed
from .models import Incident
from .filters import IncidentFilter

class ExportIncidentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as e:
            return Response(
                {"error": f"Error de dependencia: {str(e)}. Instale openpyxl en el backend."},
                status=500
            )

        # 1. Filtrado de datos (reutilizando IncidentFilter)
        queryset = Incident.objects.all().select_related(
            'created_by', 'assigned_to', 'categoria', 'responsable'
        ).order_by('-created_at')
        
        filterset = IncidentFilter(request.GET, queryset=queryset)
        if filterset.is_valid():
            queryset = filterset.qs
            
        # 2. Creación del Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Incidencias"
        
        # 3. Estilos Corporativos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Blue-600
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 4. Definición de Columnas (Más exhaustivo que la vista web)
        columns = [
            ("Código", "code", 15),
            ("Fecha Detección", "fecha_deteccion", 15),
            ("Hora", "hora_deteccion", 10),
            ("Cliente", "cliente", 25),
            ("RUT Cliente", "cliente_rut", 15),
            ("Obra", "obra", 20),
            ("Proveedor", "provider", 20),
            ("Categoría", "categoria__name", 15),
            ("Subcategoría", "subcategoria", 15),
            ("Estado", "estado", 15),
            ("Prioridad", "prioridad", 12),
            ("Descripción", "descripcion", 40),
            ("Creado por", "created_by__username", 15),
            ("Fecha Creación", "created_at", 20),
        ]
        
        # 5. Escribir Encabezados
        for col_num, (header_title, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            # Ajustar ancho de columna
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # 6. Escribir Datos
        for row_num, incident in enumerate(queryset, 2):
            for col_num, (_, field_path, _) in enumerate(columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Obtener valor (soporte para relaciones __)
                value = None
                try:
                    if '__' in field_path:
                        parts = field_path.split('__')
                        obj = incident
                        for part in parts:
                            obj = getattr(obj, part, None)
                            if obj is None: break
                        value = obj
                    else:
                        value = getattr(incident, field_path, None)
                except Exception:
                    value = ""

                # Formateo de valores
                if value is None:
                    cell.value = ""
                elif field_path == 'estado':
                    cell.value = dict(Incident.STATUS_CHOICES).get(value, value).replace('_', ' ').title()
                elif isinstance(value, (timezone.datetime, datetime.date)):
                    if isinstance(value, timezone.datetime):
                        cell.value = value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        cell.value = value.strftime('%Y-%m-%d')
                else:
                    cell.value = str(value)
                
                cell.alignment = left_align
                cell.border = thin_border

        # 7. Respuesta HTTP con archivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Incidencias_Export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
